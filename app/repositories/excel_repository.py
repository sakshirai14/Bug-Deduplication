import pandas as pd
from io import BytesIO
from typing import List, BinaryIO
import openpyxl
from openpyxl.styles import Alignment

class ExcelRepository:
    def read_excel(self, file_content: BinaryIO) -> pd.DataFrame:
        df = pd.read_excel(file_content)
        df.columns = df.columns.astype(str).str.strip()
        return df

    def append_results_to_excel(self, original_file_content: BinaryIO, results: List[dict]) -> BytesIO:
        # Load workbook
        wb = openpyxl.load_workbook(original_file_content)
        ws = wb.active
        
        # Determine header row (assuming row 1)
        # Find headers for Result, Matching IDs, Match Confidence
        # Append them if they don't exist
        
        headers = ["Result", "Matching IDs", "Match Confidence"]
        base_col_idx = ws.max_column + 1
        
        for i, header in enumerate(headers):
            ws.cell(row=1, column=base_col_idx + i, value=header)
            
        # Write results
        # results is list of dict with keys: result, matching_ids, match_confidence
        # matching rows 2 to len(results)+1
        
        for i, res in enumerate(results):
            row_idx = i + 2 # 1-based, skip header
            
            ws.cell(row=row_idx, column=base_col_idx, value=res.get("result", ""))
            
            # Matching IDs with wrap text
            cell_ids = ws.cell(row=row_idx, column=base_col_idx + 1, value=res.get("matching_ids", ""))
            cell_ids.alignment = Alignment(wrap_text=True)
            
            ws.cell(row=row_idx, column=base_col_idx + 2, value=res.get("match_confidence", ""))
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
